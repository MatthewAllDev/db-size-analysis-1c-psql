import openpyxl
import openpyxl.utils
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from .row import Row


class ExcelProcessor:
    def __init__(self, file_path: str = None):
        if file_path is None:
            self.__workbook: openpyxl.Workbook = openpyxl.Workbook()
        else:
            self.__workbook: openpyxl.Workbook = openpyxl.load_workbook(file_path)
        self.__sheet: Worksheet = self.__workbook.active
        self.__current_row_index: int = 1

    def add_row(self, *args) -> Row:
        row: Row = Row(self.__sheet, self.__current_row_index)
        row.set(*args)
        self.__current_row_index += 1
        return row

    def create_rows_tree(self, start: int,
                         end: int,
                         children_index: list = None,
                         outline_level: int = 1,
                         hidden: bool = True):
        """
        :param start: first row to be grouped
        :param end: last row to be grouped
        :param children_index: list[dict{'start': int, 'end': int, 'children_index': list (optional)}] optional
        :param outline_level: outline level (optional, default to 1)
        :param hidden: should the group be hidden on workbook open or not (optional, default to True)
        """
        self.group_rows(start, end, outline_level, hidden)
        if children_index is not None:
            for child_index in children_index:
                try:
                    range_size: int = child_index['end'] - child_index['start']
                except IndexError:
                    raise IndexError('keys "start" and "end" must be defined in "children_index"')
                if range_size > 0:
                    self.create_rows_tree(child_index['start'],
                                          child_index['end'],
                                          child_index.get('children_index'),
                                          outline_level + 1,
                                          hidden)

    def group_rows(self, start: int, end: int = None, outline_level: int = 1, hidden: bool = True):
        """
        For tree: use outline_level 1 to X in order. Outline_level overwritten for range.

        :param start: first row to be grouped (mandatory)
        :param end: last row to be grouped (optional, default to start)
        :param outline_level: outline level (optional, default to 1)
        :param hidden: should the group be hidden on workbook open or not (optional, default to True)
        """
        self.__sheet.row_dimensions.group(start, end, outline_level, hidden)

    def create_columns_tree(self, start: int,
                            end: int,
                            children_index: list = None,
                            outline_level: int = 1,
                            hidden: bool = True):
        """
        :param start: first column to be grouped
        :param end: last column to be grouped
        :param children_index: list[dict{'start': int, 'end': int, 'children_index': list (optional)}] optional
        :param outline_level: outline level (optional, default to 1)
        :param hidden: should the group be hidden on workbook open or not (optional, default to True)
        """
        self.group_columns(start, end, outline_level, hidden)
        if children_index is not None:
            for child_index in children_index:
                try:
                    range_size: int = child_index['end'] - child_index['start']
                except IndexError:
                    raise IndexError('keys "start" and "end" must be defined in "children_index"')
                if range_size > 0:
                    self.create_columns_tree(child_index['start'],
                                             child_index['end'],
                                             child_index.get('children_index'),
                                             outline_level + 1,
                                             hidden)

    def group_columns(self, start: int, end: int = None, outline_level: int = 1, hidden: bool = True):
        """
        For tree: use outline_level 1 to X in order. Outline_level overwritten for range.

        :param start: first column to be grouped (mandatory)
        :param end: last column to be grouped (optional, default to start)
        :param outline_level: outline level (optional, default to 1)
        :param hidden: should the group be hidden on workbook open or not (optional, default to True)
        """
        self.__sheet.column_dimensions.group(start, end, outline_level, hidden)

    def set_optimal_column_widths(self):
        column_widths: list = []
        for row in self.__sheet.rows:
            for index, cell in enumerate(row):
                if len(column_widths) > index:
                    if len(str(cell.value)) > column_widths[index]:
                        column_widths[index] = len(str(cell.value))
                else:
                    column_widths.append(len(str(cell.value)))
        for index, column_width in enumerate(column_widths):
            self.__sheet.column_dimensions[openpyxl.utils.get_column_letter(index + 1)].width = column_width + 1

    def save(self, file_path: str) -> str:
        self.__workbook.save(filename=file_path)
        return file_path

    def set_orientation(self, orientation: str):
        if orientation.lower() == 'landscape':
            self.__sheet.page_setup.orientation = self.__sheet.ORIENTATION_LANDSCAPE
        elif orientation.lower() == 'portrait':
            self.__sheet.page_setup.orientation = self.__sheet.ORIENTATION_PORTRAIT
        else:
            raise AttributeError('Unsupported orientation')

    def set_fit_to_width(self):
        self.set_fit_to_page()
        self.__sheet.page_setup.fitToHeight = False

    def set_fit_to_height(self):
        self.set_fit_to_page()
        self.__sheet.page_setup.fitToWidth = False

    def set_fit_to_page(self):
        self.__sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)

    def set_margins(self, left: float = None, right: float = None, top: float = None, bottom: float = None):
        if left is not None:
            self.__sheet.page_margins.left = left
        if right is not None:
            self.__sheet.page_margins.right = right
        if top is not None:
            self.__sheet.page_margins.top = top
        if bottom is not None:
            self.__sheet.page_margins.bottom = bottom
